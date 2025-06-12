import openpyxl
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

# Parameters and default values
parameters = [
    ("Villa_Price", 5500000),
    ("Upfront_Percentage", 0.27),
    ("Enhancements", 100000),
    ("LTV", 0.80),
    ("Loan_Term_Yrs", 25),
    ("Fixed_Rate", 0.0411),
    ("Fixed_Rate_Years", 3),
    ("Variable_Rate_Base", 0.047),
    ("Insurance_Monthly", 500),
    ("Insurance_Growth", 0.03),
    ("Service_Fee_Annual", 15000),
    ("Service_Fee_Growth", 0.03),
    ("Property_CAGR", 0.03),
    ("Portfolio_CAGR", 0.07),
    ("Annual_Savings", 500000),
    ("Current_Rent", 220000),
    ("Rent_Inflation", 0.03),
    ("Analysis_Years", 15),
    ("Equity_Market_Shock_Yr1", 0),
]

wb = Workbook()
# remove default sheet
wb.remove(wb.active)

# ---------------- Inputs Sheet -----------------
ws_inputs = wb.create_sheet("Inputs")
ws_inputs["A1"] = "Parameter"
ws_inputs["B1"] = "Value"

for idx, (p, v) in enumerate(parameters, start=2):
    ws_inputs.cell(row=idx, column=1, value=p)
    ws_inputs.cell(row=idx, column=2, value=v)
    ref = f"'{ws_inputs.title}'!$B${idx}"
    wb.defined_names.append(DefinedName(p, attr_text=ref))

# ---------------- Yearly Projection -----------------
ws_yearly = wb.create_sheet("Yearly_Buy_vs_Rent")
headers = [
    "Year",
    "Mortgage Payment",
    "Remaining Principal",
    "Property Value",
    "Property Equity",
    "Portfolio_Value_if_Buy",
    "Net_Worth_if_Buy",
    "Portfolio_Value_if_Rent",
    "Net_Worth_if_Rent",
    "Advantage_vs_Rent",
]
for col, h in enumerate(headers, start=1):
    ws_yearly.cell(row=1, column=col, value=h)

analysis_years = 15
for year in range(1, analysis_years + 1):
    r = year + 1
    ws_yearly.cell(row=r, column=1, value=year)
    # Mortgage Payment
    ws_yearly.cell(row=r, column=2, value=
        f"=PMT(IF(A{r}<=Fixed_Rate_Years,Fixed_Rate,Variable_Rate_Base)/12,"
        f" Loan_Term_Yrs*12-(A{r}-1)*12, -Villa_Price*LTV)")
    # Remaining Principal
    ws_yearly.cell(row=r, column=3, value=
        f"=FV(IF(A{r}<=Fixed_Rate_Years,Fixed_Rate,Variable_Rate_Base)/12,12,-B{r},"
        f" IF(A{r}=1,Villa_Price*LTV,C{r-1}))")
    # Property Value
    ws_yearly.cell(row=r, column=4, value=f"=Villa_Price*(1+Property_CAGR)^A{r}")
    # Property Equity
    ws_yearly.cell(row=r, column=5, value=f"=D{r}-C{r}")
    # Portfolio Value if Buy
    prev_port_buy = "0" if year == 1 else f"F{r-1}"
    ws_yearly.cell(row=r, column=6, value=f"=FV(Portfolio_CAGR,1,-Annual_Savings,{prev_port_buy})")
    # Net Worth if Buy
    ws_yearly.cell(row=r, column=7, value=f"=E{r}+F{r}")
    # Portfolio Value if Rent
    prev_port_rent = "0" if year == 1 else f"H{r-1}"
    ws_yearly.cell(row=r, column=8, value=f"=FV(Portfolio_CAGR,1,-Annual_Savings,{prev_port_rent})")
    # Net Worth if Rent
    ws_yearly.cell(row=r, column=9, value=f"=H{r}")
    # Advantage vs Rent
    ws_yearly.cell(row=r, column=10, value=f"=G{r}-I{r}")

# ---------------- Scenario Summary -----------------
ws_scen = wb.create_sheet("Scenario_Summary")
scen_headers = [
    "Scenario",
    "VariableRateAfterFix",
    "RentInflation",
    "PortfolioCAGR",
    "PropertyCAGR",
    "NetWorth_Buy_15Y",
    "NetWorth_Rent_15Y",
    "Advantage",
    "Breakeven_Year",
]
for col, h in enumerate(scen_headers, start=1):
    ws_scen.cell(row=1, column=col, value=h)

scenarios = [
    "Base",
    "Rate +2 pp",
    "Rate +4 pp",
    "Equity_Crash_-30%",
    "Rent_Inflation_7%",
    "Bullish",
]
for idx, name in enumerate(scenarios, start=2):
    ws_scen.cell(row=idx, column=1, value=name)
    # Variable Rate After Fix
    ws_scen.cell(row=idx, column=2,
        value=f"=IF(A{idx}='Rate +2 pp',Variable_Rate_Base+0.02,IF(A{idx}='Rate +4 pp',Variable_Rate_Base+0.04,Variable_Rate_Base))")
    # Rent Inflation
    ws_scen.cell(row=idx, column=3,
        value=f"=IF(A{idx}='Rent_Inflation_7%',0.07,Rent_Inflation)")
    # Portfolio CAGR
    ws_scen.cell(row=idx, column=4,
        value=f"=IF(A{idx}='Bullish',0.12,Portfolio_CAGR)")
    # Property CAGR
    ws_scen.cell(row=idx, column=5,
        value=f"=IF(A{idx}='Bullish',0.05,IF(A{idx}='Equity_Crash_-30%',-0.30,Property_CAGR))")
    # Net Worth Buy 15Y (base results referenced)
    ws_scen.cell(row=idx, column=6,
        value="=INDEX(Yearly_Buy_vs_Rent!G:G,Analysis_Years+1)")
    # Net Worth Rent 15Y (base results referenced)
    ws_scen.cell(row=idx, column=7,
        value="=INDEX(Yearly_Buy_vs_Rent!I:I,Analysis_Years+1)")
    # Advantage
    ws_scen.cell(row=idx, column=8, value=f"=F{idx}-G{idx}")
    # Breakeven Year
    ws_scen.cell(row=idx, column=9,
        value="=MATCH(TRUE,Yearly_Buy_vs_Rent!J2:J16>=0,0)")

wb.save("Rent_vs_Buy.xlsx")
